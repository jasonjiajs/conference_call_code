### this code serves to bold the keywords ### 

#%%%
from operator import truediv
import openpyxl
import xlsxwriter
import re

#%% This wasn't in the original code...
os.chdir("C:/Users/jasonjia/Dropbox/ConferenceCall/Output/KeywordIdentification")

#%%%
raw_data = openpyxl.load_workbook(r"cric1_newtotal_countryadded_interestrateattheback.xlsx")
raw_data = raw_data[raw_data.sheetnames[0]] # choose first sheet of the workbook 
# (there's only one sheet but we want a sheet object)

#%%
row = 0 
col = 0
with open("keyterms.txt", "r", encoding="utf-8", errors="ignore") as f1:
    key_set = set(f1.read().splitlines())
keyw_list = list(key_set) 
keyw_list = [t.lower() for t in keyw_list]


#%%% title
title_file = openpyxl.load_workbook(r"Entry mask.xlsx")
title_file = title_file[title_file.sheetnames[0]]

#%%% overview
overview_file = xlsxwriter.Workbook("Paragraph Record.xlsx")
w1 = overview_file.add_worksheet()
w1.write(0,0,"File Name")
w1.write(0,1,"Starting Row")
w1.write(0,2,"Ending Row")
w1.write(0,3,"Name of RA")

dict_title = {0:2, 2:6, 3:7, 4:8, 5:9, 13:44, 12:45} #key = col from entry mask, value = col from cric1_newtotal

col_width = [13.29, 14.43, 30.43,  8.29,  8.29, 84.29, 13.57, 13.71, 61.57,    35,
             12.29, 11.43, 11.43, 13.43, 13.29, 10.71, 11.86, 10.86, 12.86, 18.29, 
                34, 15.71, 19.71, 14.29, 22.14, 17.29, 29.14, 10.14,  8.29,  8.29,
              9.71,  8.29,  8.86, 16.71, 15.86, 21.86, 17.29, 21.57,    20,    17, 
             24.86,    21,  9.29,    10, 14.57,    19] #from template given
row_height = 15 #

#%%
for row_val in range(1,raw_data.max_row):
    if row_val%500==1:
        if row_val//500>0:
            final_data.close()
        final_data = xlsxwriter.Workbook(r"entry_files/"+ str(row_val//500+1) + ".xlsx")
        bold_1 = final_data.add_format({"bold":True, "align":"center","valign":"center"}) # the bold "format"

        cell_format = final_data.add_format()
        cell_format.set_align('center') #horizontal center
        cell_format.set_align('vcenter') #vertical center
        cell_format.set_text_wrap()

        #cell_leftalign = final_data.add_format()
        #cell_format.set_align('left') #horizontal center
        #cell_format.set_align('vcenter') #vertical center
        #cell_format.set_text_wrap()

        #cellborderright = final_data.add_format()
        #cellborderright.set_border()
        #cellborderright.set_bottom(2)
        #cellborderright.set_left(0)
        #cellborderright.set_top(0)
        #cellborderright.set_border_color('black')
        #cellborderbottom = final_data.add_format()
        #cellborderbottom.set_bottom(1)

        worksheet = final_data.add_worksheet()

        #[worksheet.set_column(i, 5, cellborderright) for i in ['C:C']]
        #worksheet.set_column(3,3, cellborderright)
        for i in range(46):
            #if i != 5:
            worksheet.set_column(i, i, col_width[i],cell_format)
            #else:
                #worksheet.set_column(i, i, col_width[i],cell_leftalign)
        for i in range(2,502):
            worksheet.set_row(i, row_height,cell_format)
        #[worksheet.set_column(i, i, col_width[i],cellborderright) for key, i in enumerate(['C', 'E', 'F', 'J', 'M', 'P', 'S', 'AA', 'AD', 'AG', 'AP', 'AR'])]
        #worksheet.set_row(i, 15, cellborderbottom)

        w1.write(row_val//500+1,0,row_val//500+1)
        w1.write(row_val//500+1,1,row_val)
        if row_val//500 + 1>1:
            w1.write(row_val//500,2,row_val-1)//7
        #%%% write title 
        for row_val_1 in range(2):
            for col_val_1 in range(title_file.max_column):
                worksheet.write(row_val_1,col_val_1,title_file.cell(row_val_1+1,col_val_1+1).value,bold_1) # write out the titles

    keyw = raw_data.cell(row_val+1,0+1).value
    text = raw_data.cell(row_val+1,1+1).value
    in_list_key = []
    for i in keyw_list:
        if i in text.lower():
            in_list_key.append(i)
    new_text = re.split("|".join(in_list_key),text,flags=re.I)
    format_text = []
    for i in range(len(new_text)):
        splited_text = new_text[i]
        if splited_text!="":
            format_text.append(splited_text)
        if i<len(new_text)-1:
            format_text.append(bold_1)
            for each_key in in_list_key:
                test_text = splited_text + each_key
                if test_text.lower() in text.lower():
                    format_text.append(each_key)
                    break
    for col_val in dict_title.keys():
        worksheet.write((row_val-1)%500+2, dict_title[col_val],raw_data.cell(row_val+1,col_val+1).value)

    worksheet.write_rich_string((row_val-1)%500+2,5,*format_text)
    worksheet.write((row_val-1)%500+2, 1, row_val)
    if row_val == raw_data.max_row-1:
        w1.write(row_val//500+1,2,row_val)
        final_data.close()


overview_file.close()
# %%

# %%

# %%

# %%
